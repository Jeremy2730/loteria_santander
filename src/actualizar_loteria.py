"""
=============================================================
  ACTUALIZADOR AUTOMÁTICO — LOTERÍA DE SANTANDER
  Fuente: astroluna.co/santander
  Ejecutar cada viernes después de las 11:00 PM
=============================================================
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import os
import sys
import logging
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(ROOT_DIR)
from config import EXCEL_FILE, LOGS_DIR
# ── CONFIGURACIÓN ─────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)

RUTA_EXCEL = EXCEL_FILE
LOG_FILE   = os.path.join(LOGS_DIR, "log_actualizacion.txt")

URL_FUENTE = "https://www.astroluna.co/santander"

# ──────────────────────────────────────────────────────────

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
console = logging.StreamHandler(sys.stdout)
console.setLevel(logging.INFO)
logging.getLogger().addHandler(console)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def obtener_ultimo_resultado():
    """Descarga la página y extrae el resultado más reciente."""
    logging.info("Consultando %s ...", URL_FUENTE)
    try:
        resp = requests.get(URL_FUENTE, timeout=15,
                            headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
    except Exception as e:
        logging.error("Error al descargar la página: %s", e)
        return None

    soup = BeautifulSoup(resp.text, "html.parser")

    # La tabla de resultados tiene encabezados Fecha / Número / Serie
    tabla = soup.find("table")
    if not tabla:
        logging.error("No se encontró tabla de resultados en la página.")
        return None

    filas = tabla.find_all("tr")
    for fila in filas[1:]:  # saltar cabecera
        celdas = [td.get_text(strip=True) for td in fila.find_all("td")]
        if len(celdas) >= 3:
            fecha_texto = celdas[0]   # ej. "viernes 06 marzo 2026"
            numero      = celdas[1]   # ej. "5904"
            serie       = celdas[2]   # ej. "034"

            # Convertir "viernes 06 marzo 2026" → "2026-03-06"
            meses = {
                "enero":"01","febrero":"02","marzo":"03","abril":"04",
                "mayo":"05","junio":"06","julio":"07","agosto":"08",
                "septiembre":"09","octubre":"10","noviembre":"11","diciembre":"12"
            }
            partes = fecha_texto.lower().split()
            try:
                # formato: [dia_semana, dia, mes, año]
                dia  = partes[1].zfill(2)
                mes  = meses.get(partes[2], "00")
                anio = partes[3]
                fecha_iso = f"{anio}-{mes}-{dia}"
            except Exception:
                logging.warning("No se pudo parsear fecha: '%s'", fecha_texto)
                continue

            if len(numero) == 4 and numero.isdigit():
                return {"fecha": fecha_iso, "numero": numero, "serie": serie}

    logging.error("No se pudo extraer ningún resultado válido.")
    return None

def ya_existe_en_excel(ws, fecha):
    """Revisa si la fecha ya está registrada en la hoja Histórico."""
    for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
        if row[0] == fecha:
            return True
    return False

def agregar_resultado(resultado):
    """Inserta el nuevo resultado en la fila 5 (debajo del encabezado) y desplaza los demás."""
    if not os.path.exists(RUTA_EXCEL):
        logging.error("No se encontró el archivo Excel en: %s", RUTA_EXCEL)
        logging.error("Verifica que RUTA_EXCEL apunte al archivo correcto.")
        return False

    wb = load_workbook(RUTA_EXCEL)

    if "Histórico" not in wb.sheetnames:
        logging.error("La hoja 'Histórico' no existe en el Excel.")
        return False

    ws = wb["Histórico"]
    fecha  = resultado["fecha"]
    numero = resultado["numero"]
    serie  = resultado["serie"]

    if ya_existe_en_excel(ws, fecha):
        logging.info("El resultado del %s ya está registrado. No se hace nada.", fecha)
        return False

    # Calcular número de sorteo (el de la fila 5 actual + 1)
    sorteo_anterior = ws["B5"].value or 0
    nuevo_sorteo    = (sorteo_anterior + 1) if isinstance(sorteo_anterior, int) else sorteo_anterior

    # Insertar fila nueva en la posición 5 (empuja todo hacia abajo)
    ws.insert_rows(5)

    c_dark  = "1F4E79"
    c_light = "DEEAF1"

    valores = [fecha, nuevo_sorteo, numero, serie,
               numero[0], numero[1], numero[2], numero[3]]
    cols = ["A","B","C","D","E","F","G","H"]

    for col, val in zip(cols, valores):
        cell = ws[f"{col}5"]
        cell.value = val
        cell.fill      = PatternFill("solid", start_color=c_light)
        cell.font      = Font(
            bold   = (col == "C"),
            name   = "Arial",
            size   = 10,
            color  = c_dark if col == "C" else "000000"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
    ws.row_dimensions[5].height = 15

    wb.save(RUTA_EXCEL)
    logging.info("✅  Resultado agregado: %s | %s | Serie %s", fecha, numero, serie)
    return True

def main():
    logging.info("=" * 55)
    logging.info("Iniciando actualización — %s", datetime.now().strftime("%d/%m/%Y %H:%M"))

    resultado = obtener_ultimo_resultado()
    if resultado:
        logging.info("Resultado encontrado: %s — %s (serie %s)",
                     resultado["fecha"], resultado["numero"], resultado["serie"])
        agregar_resultado(resultado)
    else:
        logging.warning("No se pudo obtener el resultado. Intenta de nuevo más tarde.")

    logging.info("Proceso finalizado.")
    logging.info("=" * 55)

if __name__ == "__main__":
    main()
